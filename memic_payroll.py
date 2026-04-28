#!/usr/bin/env python3
"""Generate MEMIC Comp-As-You-Go payroll submission files from Square payroll CSV exports."""

import csv
import hashlib
import json
import os
from difflib import get_close_matches
from pathlib import Path

import openpyxl
from openpyxl import Workbook


EMPLOYEES_FILE = Path(__file__).parent / "employees.json"
OUTPUT_DIR = Path(__file__).parent / "output"

MEMIC_HEADERS = [
    "Employee Name",
    "Employee ID",
    "State - Class Code",
    "Gross Wages - Includes OT, DT and Tips",
    "Overtime - Including Time and a Half",
    "Tips - as included in Gross Wages",
    "Double Time - Including Time and Double Time Pay",
    "Section 125",
    "Hours",
]


def load_db() -> dict:
    if EMPLOYEES_FILE.exists():
        with open(EMPLOYEES_FILE) as f:
            return json.load(f)
    return {"square_name_map": {}, "memic_roster": []}


def save_db(db: dict):
    with open(EMPLOYEES_FILE, "w") as f:
        json.dump(db, f, indent=2)


def parse_square_csv(path: str) -> tuple[dict, str, str, str]:
    with open(path, newline="") as f:
        rows = list(csv.reader(f))

    pay_period_start = rows[0][1].strip()
    pay_period_end = rows[1][1].strip()
    pay_date = rows[2][1].strip()

    header = [h.strip() for h in rows[4]]
    col = {name: i for i, name in enumerate(header)}

    def flt(row, name) -> float:
        idx = col.get(name)
        if idx is None or idx >= len(row):
            return 0.0
        try:
            return float(row[idx])
        except ValueError:
            return 0.0

    employees = {}
    for row in rows[5:]:
        if not row or row[0].strip() in ("Total", ""):
            break
        first = row[col["First Name"]].strip()
        last = row[col["Last Name"]].strip()
        employees[f"{first} {last}"] = {
            "gross": flt(row, "Gross Pay"),
            "overtime": flt(row, "OT Earnings"),
            "double_time": flt(row, "DT Earnings"),
            "tips": flt(row, "Paycheck Tips"),
            "hours": (flt(row, "Reg Hours") + flt(row, "OT Hours") +
                      flt(row, "DT Hours") + flt(row, "PTO Hours")),
        }

    return employees, pay_period_start, pay_period_end, pay_date


def parse_memic_template(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb["Form"]
    roster = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        if not row[0]:
            continue
        roster.append({
            "name": str(row[0]).strip(),
            "id": str(row[1]).strip() if row[1] else "",
            "class_code": str(row[2]).strip() if row[2] else "",
        })
    return roster


def auto_match(square_name: str, roster: list[dict]) -> str | None:
    """Return the best MEMIC name for a Square employee name, or None."""
    memic_names = [e["name"] for e in roster]

    parts = square_name.split(maxsplit=1)
    if len(parts) == 2:
        last_first = f"{parts[1]}, {parts[0]}"

        # Exact "Last, First" match
        if last_first in memic_names:
            return last_first

        # Last name only (handles "Valerie" → "Val" first-name mismatches)
        last = parts[1].lower()
        by_last = [n for n in memic_names if n.split(",")[0].strip().lower() == last]
        if len(by_last) == 1:
            return by_last[0]

        # Fuzzy fallback
        matches = get_close_matches(last_first, memic_names, n=1, cutoff=0.6)
        if matches:
            return matches[0]

    return None


def update_roster(db: dict, template_roster: list[dict]):
    """Merge template employees into the stored roster, updating changed records."""
    existing = {e["name"]: e for e in db["memic_roster"]}
    for emp in template_roster:
        if emp["name"] in existing:
            existing[emp["name"]].update(emp)
        else:
            db["memic_roster"].append(emp)
            existing[emp["name"]] = emp


def generate_xlsx(db: dict, payroll: dict, pay_date: str) -> Path:
    memic_payroll = {}
    for sq_name, data in payroll.items():
        memic_name = db["square_name_map"].get(sq_name)
        if memic_name:
            memic_payroll[memic_name] = data

    wb = Workbook()
    ws = wb.active
    ws.title = "Form"
    ws.append(MEMIC_HEADERS)

    for emp in db["memic_roster"]:
        data = memic_payroll.get(emp["name"])
        if data:
            ws.append([
                emp["name"], emp["id"], emp["class_code"],
                data["gross"], data["overtime"], data["tips"],
                data["double_time"], 0, data["hours"],
            ])
        else:
            ws.append([emp["name"], emp["id"], emp["class_code"], 0, 0, 0, 0, 0, 0])

    OUTPUT_DIR.mkdir(exist_ok=True)
    parts = pay_date.split("/")
    date_str = f"{parts[2]}-{parts[0]}-{parts[1]}" if len(parts) == 3 else pay_date.replace("/", "-")
    fingerprint = hashlib.md5(
        json.dumps(sorted(payroll.items()), sort_keys=True).encode()
    ).hexdigest()[:8]
    out_path = OUTPUT_DIR / f"MEMIC_Payroll_{date_str}_{fingerprint}.xlsx"
    wb.save(out_path)
    return out_path


def prompt_path(prompt_text: str) -> str:
    while True:
        raw = input(prompt_text).strip().strip("'\"")
        path = os.path.expanduser(raw)
        if os.path.exists(path):
            return path
        print(f"  File not found: {path}")


def resolve_match(sq_name: str, memic_name: str, db: dict, roster: list[dict]):
    """Confirm or manually select a MEMIC match for a Square employee name."""
    parts = sq_name.split(maxsplit=1)
    last_first = f"{parts[1]}, {parts[0]}" if len(parts) == 2 else sq_name
    tag = "exact" if memic_name == last_first else "fuzzy"
    answer = input(f"  Match '{sq_name}' → '{memic_name}' ({tag})? [Y/n]: ").strip().lower()
    if answer in ("", "y", "yes"):
        db["square_name_map"][sq_name] = memic_name
        return

    memic_names = [e["name"] for e in roster]
    print("  Available MEMIC names:")
    for i, n in enumerate(memic_names, 1):
        print(f"    {i}. {n}")
    choice = input(f"  Enter number or name (blank to skip): ").strip()
    if choice.isdigit() and 1 <= int(choice) <= len(memic_names):
        db["square_name_map"][sq_name] = memic_names[int(choice) - 1]
    elif choice in memic_names:
        db["square_name_map"][sq_name] = choice


def main():
    print("MEMIC Payroll Generator\n")

    # Step 1: Square payroll CSV
    csv_path = prompt_path("Square payroll CSV path: ")
    payroll, period_start, period_end, pay_date = parse_square_csv(csv_path)
    print(f"  Pay period: {period_start} – {period_end}  |  Pay date: {pay_date}")
    print(f"  Employees in run: {', '.join(payroll.keys())}\n")

    db = load_db()

    # Bucket employees: already mapped / matchable against roster / truly unknown
    already_mapped = [n for n in payroll if n in db["square_name_map"]]
    needs_match = []   # (sq_name, memic_name) — auto-matched against stored roster
    truly_unknown = [] # sq_names with no match in stored roster at all

    for sq_name in payroll:
        if sq_name in db["square_name_map"]:
            continue
        memic_name = auto_match(sq_name, db["memic_roster"])
        if memic_name:
            needs_match.append((sq_name, memic_name))
        else:
            truly_unknown.append(sq_name)

    # Confirm auto-matches against stored roster (no template needed)
    if needs_match:
        print("  New employees to confirm:")
        for sq_name, memic_name in needs_match:
            resolve_match(sq_name, memic_name, db, db["memic_roster"])
        save_db(db)
        print()

    # Truly unknown — can't match against roster; need a fresh template
    if truly_unknown:
        print(f"  Not found in roster: {', '.join(truly_unknown)}")
        print("  Reminder: if any are new hires, add them to MEMIC and download")
        print("  a fresh template first — otherwise they won't appear in the file.\n")

        template_path = prompt_path("MEMIC template (.xlsx) path: ")
        template_roster = parse_memic_template(template_path)
        update_roster(db, template_roster)

        for sq_name in truly_unknown:
            memic_name = auto_match(sq_name, template_roster)
            if memic_name:
                resolve_match(sq_name, memic_name, db, template_roster)
            else:
                memic_names = [e["name"] for e in template_roster]
                print(f"  No match for '{sq_name}'. Available MEMIC names:")
                for i, n in enumerate(memic_names, 1):
                    print(f"    {i}. {n}")
                choice = input(f"  Enter number or name (blank to skip): ").strip()
                if choice.isdigit() and 1 <= int(choice) <= len(memic_names):
                    db["square_name_map"][sq_name] = memic_names[int(choice) - 1]
                elif choice in memic_names:
                    db["square_name_map"][sq_name] = choice

        save_db(db)
        print("  Employee map saved.\n")

    if not db["memic_roster"]:
        print("  Error: no MEMIC roster found. Run again with a MEMIC template to initialize.")
        return

    # Step 3: Generate output
    out_path = generate_xlsx(db, payroll, pay_date)

    print(f"Output:     {out_path}")
    print(f"Check Date: {pay_date}")


if __name__ == "__main__":
    main()
